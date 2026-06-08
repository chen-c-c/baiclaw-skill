#!/usr/bin/env python3
"""
Excel Formula Verifier — pure Python, no LibreOffice required.

Checks an Excel file for formula errors using openpyxl.
Optionally uses the `formulas` library (pip install formulas) for in-memory
formula evaluation before scanning. When `formulas` is available it writes
calculated values back to the workbook so subsequent opens in Excel show
correct values without needing manual recalculation.

Usage:
  python recalc.py <excel_file>

Returns JSON:
  {
    "status": "success" | "errors_found",
    "total_formulas": N,
    "total_errors": N,
    "engine": "formulas" | "openpyxl-scan",
    "error_summary": {
      "#REF!": {"count": 1, "locations": ["Sheet1!A3"]},
      ...
    }
  }
"""

import json
import sys
from pathlib import Path

EXCEL_ERRORS = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']


def _count_formulas(wb) -> int:
    total = 0
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    total += 1
    return total


def _scan_errors(wb) -> tuple:
    """Scan cached/calculated cell values for Excel error strings."""
    error_details = {e: [] for e in EXCEL_ERRORS}
    total = 0
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    for err in EXCEL_ERRORS:
                        if err in cell.value:
                            error_details[err].append(f"{sn}!{cell.coordinate}")
                            total += 1
                            break
    return error_details, total


def check_errors(filename):
    if not Path(filename).exists():
        return {"error": f"File '{filename}' does not exist"}

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "openpyxl not installed — run: pip install openpyxl"}

    abs_path = str(Path(filename).resolve())
    engine = "openpyxl-scan"

    # --- Attempt in-memory formula evaluation with the `formulas` library ---
    try:
        import formulas  # type: ignore
        xl = formulas.ExcelModel().loads(abs_path).finish()
        xl.calculate()
        # Write calculated values back so openpyxl data_only scan reflects them.
        xl.to_xlsx(abs_path)
        engine = "formulas"
    except ImportError:
        pass  # formulas not installed; fall back to cached-value scan
    except Exception:
        pass  # evaluation failed (unsupported functions etc.); continue with scan

    # --- Scan for error strings in (cached) cell values ---
    try:
        wb_data = load_workbook(abs_path, data_only=True)
        error_details, total_errors = _scan_errors(wb_data)
        wb_data.close()

        wb_src = load_workbook(abs_path, data_only=False)
        formula_count = _count_formulas(wb_src)
        wb_src.close()
    except Exception as exc:
        return {"error": str(exc)}

    result = {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_formulas": formula_count,
        "total_errors": total_errors,
        "engine": engine,
        "error_summary": {},
    }
    for err_type, locations in error_details.items():
        if locations:
            result["error_summary"][err_type] = {
                "count": len(locations),
                "locations": locations[:20],
            }
    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file>")
        print()
        print("Checks an Excel file for formula errors (no LibreOffice required).")
        print("Install 'formulas' for full in-memory evaluation: pip install formulas")
        sys.exit(1)

    result = check_errors(sys.argv[1])
    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
